import openpyxl
#from openpyxl.styles import Font

#Find the first empty row in column j
def findEmptyRow(dest,j):
    empty_row = 1
    while dest.cell(row=empty_row, column=j).value is not None:
        empty_row += 1
    return empty_row

#Write the content of a row from src to dest
def write(src,x,y,dest,i,j):
    toCopy = src.cell(row=x,column=y).value
    dest.cell(row=i,column=j).value = toCopy

# Copy data from Spoke5-publications to PE13 - Data for KPI
# Load the first workbook
wb1 = openpyxl.load_workbook('Spoke5-publications.xlsx')
sheet1 = wb1['papers']
# Load the second workbook
wb2 = openpyxl.load_workbook('PE13 - Data for KPI.xlsx')
sheet2 = wb2['Paper in conference']
sheet3 = wb2['Publishing']
sheet4 = wb2['Book&Book chapter']

numero_righe_file_spoke5=findEmptyRow(sheet1,1)
#per ora vado in append sugli sheets dell'excel2, ma posso scriverli anche da 0 o anche andando a specificare un qualsiasi indice.
index_conference=findEmptyRow(sheet2,2)
index_publishing=findEmptyRow(sheet3,2)
index_other=findEmptyRow(sheet4,2)
print("wb1,",str(numero_righe_file_spoke5-1)+" righe")
print("wb2 Paper in conference,","inizio a scrivere dalla riga "+str(index_conference))
print("wb2 Publishing,","inizio a scrivere dalla riga "+str(index_publishing))
print("wb2 Book&Book chapter,","inizio a scrivere dalla riga "+str(index_other))

for ith_row in range(2,numero_righe_file_spoke5):
    cell_value = sheet1.cell(row=ith_row,column=4).value #read the column Type
    if cell_value == "Journal":
        write(sheet1,ith_row,1,sheet3,index_publishing,2) #authors
        write(sheet1,ith_row,2,sheet3,index_publishing,3) #affiliation of authors
        write(sheet1,ith_row,3,sheet3,index_publishing,5) #title
        write(sheet1,ith_row,5,sheet3,index_publishing,6) #journal
        write(sheet1,ith_row,6,sheet3,index_publishing,9) #year
        write(sheet1,ith_row,7,sheet3,index_publishing,10) #doi
        write(sheet1,ith_row,8,sheet3,index_publishing,11) #link paper
        write(sheet1,ith_row,10,sheet3,index_publishing,12) #TP
        write(sheet1,ith_row,16,sheet3,index_publishing,13) #authors from other spokes
        write(sheet1,ith_row,18,sheet3,index_publishing,14) #authors from foreign affiliations
        write(sheet1,ith_row,19,sheet3,index_publishing,15) #industrial co-authorship
        write(sheet1,ith_row,20,sheet3,index_publishing,16) #topic
        write(sheet1,ith_row,21,sheet3,index_publishing,17) #multisciplinary
        write(sheet1,ith_row,24,sheet3,index_publishing,19) #status
        write(sheet1,ith_row,27,sheet3,index_publishing,7) #rank => questo copia la macro e non il risultato della macro.
        index_publishing=index_publishing+1
    elif cell_value == "Conference":
        index=findEmptyRow(sheet2,2)
        write(sheet1,ith_row,1,sheet2,index_conference,2) #authors
        write(sheet1,ith_row,2,sheet2,index_conference,4) #affiliation of authors
        write(sheet1,ith_row,3,sheet2,index_conference,3) #title
        write(sheet1,ith_row,5,sheet2,index_conference,6) #conference
        write(sheet1,ith_row,6,sheet2,index_conference,9) #year
        write(sheet1,ith_row,7,sheet2,index_conference,10) #doi
        write(sheet1,ith_row,8,sheet2,index_conference,11) #link paper
        write(sheet1,ith_row,10,sheet2,index_conference,12) #TP
        write(sheet1,ith_row,16,sheet2,index_conference,13) #authors from other spokes
        write(sheet1,ith_row,18,sheet2,index_conference,14) #authors from foreign affiliations
        write(sheet1,ith_row,19,sheet2,index_conference,15) #industrial co-authorship
        write(sheet1,ith_row,20,sheet2,index_conference,16) #topic
        write(sheet1,ith_row,21,sheet2,index_conference,17) #multisciplinary
        write(sheet1,ith_row,24,sheet2,index_conference,18) #status
        write(sheet1,ith_row,27,sheet2,index_conference,7) #rank => questo copia la macro e non il risultato della macro.
        index_conference=index_conference+1
    else: #Workshop/Other to be handled. Non capisco se nel file Book&Book Chapter ci vanno anche gli articoli di workshop. Ad esempio, vedi riga 4 dell'excel 2 in Book&Book Chapter
        pass

# Save the changes to the second workbook
wb2.save('PE13 - Data for KPI.xlsx')
